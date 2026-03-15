
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import threading
import time
import requests
import pandas as pd
import numpy as np
import matplotlib
matplotlib.use("TkAgg")
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from matplotlib.figure import Figure
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter
from datetime import datetime

# ──────────────────────────────────── THEME ────────────────────────────────
BG_DARK    = "#0D0F1A"
BG_PANEL   = "#151828"
BG_CARD    = "#1E2235"
BG_ROW_ALT = "#181B2C"
ACCENT     = "#38003C"
ACCENT2    = "#00FF85"
TEXT_MAIN  = "#EAEAEA"
TEXT_DIM   = "#7C7F9E"
TEXT_GOLD  = "#FFD700"
CL_COLOR   = "#1A73E8"
EL_COLOR   = "#FF9500"
REL_COLOR  = "#E53935"
MID_COLOR  = "#3A3F5C"

# Sort-viz colours
COL_DEFAULT = "#5C6BC0"
COL_COMPARE = "#FFC107"
COL_SWAP    = "#F44336"
COL_SORTED  = "#4CAF50"
COL_PIVOT   = "#FF9800"

FONT_TITLE = ("Segoe UI", 16, "bold")
FONT_HEAD  = ("Segoe UI", 11, "bold")
FONT_BODY  = ("Segoe UI", 10)
FONT_SMALL = ("Segoe UI", 9)
FONT_MONO  = ("Consolas", 10)

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                  "AppleWebKit/537.36 (KHTML, like Gecko) "
                  "Chrome/122.0.0.0 Safari/537.36",
    "Accept-Language": "en-US,en;q=0.9",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
}

# ═══════════════════════════ SORTING ALGORITHMS ════════════════════════════
def bubble_sort(arr):
    a = arr.copy(); n = len(a)
    for i in range(n):
        for j in range(n - i - 1):
            yield a.copy(), [j, j+1], "compare", 1, 0
            if a[j] > a[j+1]:
                a[j], a[j+1] = a[j+1], a[j]
                yield a.copy(), [j, j+1], "swap", 0, 1
        yield a.copy(), [n-1-i], "sorted_mark", 0, 0
    yield a.copy(), list(range(n)), "done", 0, 0

def selection_sort(arr):
    a = arr.copy(); n = len(a)
    for i in range(n):
        min_idx = i
        for j in range(i+1, n):
            yield a.copy(), [min_idx, j], "compare", 1, 0
            if a[j] < a[min_idx]: min_idx = j
        if min_idx != i:
            a[i], a[min_idx] = a[min_idx], a[i]
            yield a.copy(), [i, min_idx], "swap", 0, 1
        yield a.copy(), [i], "sorted_mark", 0, 0
    yield a.copy(), list(range(n)), "done", 0, 0

def insertion_sort(arr):
    a = arr.copy(); n = len(a)
    for i in range(1, n):
        key = a[i]; j = i - 1
        while j >= 0:
            yield a.copy(), [j, j+1], "compare", 1, 0
            if a[j] > key:
                a[j+1] = a[j]; j -= 1
                yield a.copy(), [j+1], "swap", 0, 1
            else:
                break
        a[j+1] = key
        yield a.copy(), [j+1], "sorted_mark", 0, 0
    yield a.copy(), list(range(n)), "done", 0, 0

def merge_sort(arr):
    a = arr.copy()
    yield from _merge_helper(a, 0, len(a)-1)
    yield a.copy(), list(range(len(a))), "done", 0, 0

def _merge_helper(a, l, r):
    if l >= r: return
    mid = (l+r)//2
    yield from _merge_helper(a, l, mid)
    yield from _merge_helper(a, mid+1, r)
    yield from _merge(a, l, mid, r)

def _merge(a, l, mid, r):
    left = a[l:mid+1].copy(); right = a[mid+1:r+1].copy()
    i = j = 0; k = l
    while i < len(left) and j < len(right):
        yield a.copy(), [l+i, mid+1+j], "compare", 1, 0
        if left[i] <= right[j]: a[k] = left[i]; i += 1
        else:                   a[k] = right[j]; j += 1
        yield a.copy(), [k], "swap", 0, 1; k += 1
    while i < len(left):
        a[k] = left[i]; i += 1; k += 1
        yield a.copy(), [k-1], "swap", 0, 1
    while j < len(right):
        a[k] = right[j]; j += 1; k += 1
        yield a.copy(), [k-1], "swap", 0, 1

def quick_sort(arr):
    a = arr.copy()
    yield from _quick_helper(a, 0, len(a)-1)
    yield a.copy(), list(range(len(a))), "done", 0, 0

def _quick_helper(a, low, high):
    if low < high:
        yield from _partition(a, low, high)
        temp = a.copy(); pivot = temp[high]; i = low-1
        for j in range(low, high):
            if temp[j] <= pivot: i += 1; temp[i], temp[j] = temp[j], temp[i]
        temp[i+1], temp[high] = temp[high], temp[i+1]; pi = i+1
        yield from _quick_helper(a, low, pi-1)
        yield from _quick_helper(a, pi+1, high)

def _partition(a, low, high):
    pivot = a[high]; i = low-1
    for j in range(low, high):
        yield a.copy(), [j, high], "compare", 1, 0
        if a[j] <= pivot:
            i += 1; a[i], a[j] = a[j], a[i]
            yield a.copy(), [i, j], "swap", 0, 1
    a[i+1], a[high] = a[high], a[i+1]
    yield a.copy(), [i+1], "sorted_mark", 0, 0

def heap_sort(arr):
    a = arr.copy(); n = len(a)
    for i in range(n//2-1, -1, -1): yield from _heapify(a, n, i)
    for i in range(n-1, 0, -1):
        a[0], a[i] = a[i], a[0]
        yield a.copy(), [0, i], "swap", 0, 1
        yield from _heapify(a, i, 0)
    yield a.copy(), list(range(n)), "done", 0, 0

def _heapify(a, n, i):
    largest = i; l, r = 2*i+1, 2*i+2
    if l < n:
        yield a.copy(), [largest, l], "compare", 1, 0
        if a[l] > a[largest]: largest = l
    if r < n:
        yield a.copy(), [largest, r], "compare", 1, 0
        if a[r] > a[largest]: largest = r
    if largest != i:
        a[i], a[largest] = a[largest], a[i]
        yield a.copy(), [i, largest], "swap", 0, 1
        yield from _heapify(a, n, largest)

ALGORITHMS = {
    "Bubble Sort":    bubble_sort,
    "Selection Sort": selection_sort,
    "Insertion Sort": insertion_sort,
    "Merge Sort":     merge_sort,
    "Quick Sort":     quick_sort,
    "Heap Sort":      heap_sort,
}

COMPLEXITY = {
    "Bubble Sort":    ("O(n)",       "O(n²)",       "O(1)"),
    "Selection Sort": ("O(n²)",      "O(n²)",       "O(1)"),
    "Insertion Sort": ("O(n)",       "O(n²)",       "O(1)"),
    "Merge Sort":     ("O(n log n)", "O(n log n)",  "O(n)"),
    "Quick Sort":     ("O(n log n)", "O(n²)",       "O(log n)"),
    "Heap Sort":      ("O(n log n)", "O(n log n)",  "O(1)"),
}

# ════════════════════════════ DATA FETCHER ═════════════════════════════════
class DataFetcher:
    COLS       = ["Rank","Team","MP","W","D","L","GF","GA","GD","Pts"]
    EXTRA_COLS = ["xG","xGA"]

    def fetch(self):
        for fn, src in [(self._from_fbref, "fbref.com"),
                        (self._from_wikipedia, "Wikipedia")]:
            try:
                df, s = fn()
                if df is not None and len(df) >= 18:
                    return df, s
            except Exception:
                pass
        return self._sample_data(), "Sample (offline)"

    def _from_fbref(self):
        url = "https://fbref.com/en/comps/9/Premier-League-Stats"
        r = requests.get(url, headers=HEADERS, timeout=15); r.raise_for_status()
        tables = pd.read_html(r.text, flavor="lxml")
        df = tables[0].copy()
        df.columns = [c[1] if isinstance(c, tuple) else c for c in df.columns]
        rename = {"Squad":"Team","Rk":"Rank"}
        df = df.rename(columns=rename)
        for col in self.EXTRA_COLS:
            if col not in df.columns: df[col] = np.nan
        wanted = self.COLS + [c for c in self.EXTRA_COLS if c in df.columns]
        df = df[[c for c in wanted if c in df.columns]]
        return self._clean(df), "fbref.com"

    def _from_wikipedia(self):
        url = "https://en.wikipedia.org/wiki/2024%E2%80%9325_Premier_League"
        r = requests.get(url, headers=HEADERS, timeout=15); r.raise_for_status()
        tables = pd.read_html(r.text, flavor="lxml")
        for tbl in tables:
            cols = [str(c).lower() for c in tbl.columns]
            if "pts" in cols and ("gf" in cols or "f" in cols):
                df = tbl.copy(); df.columns = [str(c) for c in df.columns]
                c_map = {}
                for col in df.columns:
                    cl = col.lower()
                    if cl in ("pos","rk","#"):           c_map[col]="Rank"
                    elif cl in ("squad","team","club"):  c_map[col]="Team"
                    elif cl=="mp":  c_map[col]="MP"
                    elif cl=="w":   c_map[col]="W"
                    elif cl=="d":   c_map[col]="D"
                    elif cl=="l":   c_map[col]="L"
                    elif cl in ("gf","f"):  c_map[col]="GF"
                    elif cl in ("ga","a"):  c_map[col]="GA"
                    elif cl=="gd":  c_map[col]="GD"
                    elif cl=="pts": c_map[col]="Pts"
                df = df.rename(columns=c_map)
                for col in self.EXTRA_COLS:
                    if col not in df.columns: df[col]=np.nan
                if "Rank" not in df.columns: df.insert(0,"Rank",range(1,len(df)+1))
                wanted = self.COLS + self.EXTRA_COLS
                df = df[[c for c in wanted if c in df.columns]]
                df = self._clean(df)
                if len(df) >= 18: return df, "Wikipedia"
        return None, ""

    def _clean(self, df):
        num = ["Rank","MP","W","D","L","GF","GA","GD","Pts","xG","xGA"]
        for c in num:
            if c in df.columns: df[c] = pd.to_numeric(df[c], errors="coerce")
        df = df.dropna(subset=["Team","Pts"])
        df = df[~df["Team"].str.contains("Squad|Total|team", na=False, case=False)]
        if "Rank" not in df.columns or df["Rank"].isna().all():
            df["Rank"] = range(1, len(df)+1)
        df = df.reset_index(drop=True)
        df["Rank"] = df["Rank"].astype(int)
        return df

    def _sample_data(self):
        data = [
            (1,"Liverpool",       29,21,6,2, 62,27,35,69,57.3,29.1),
            (2,"Arsenal",         29,18,7,4, 58,29,29,61,53.0,31.0),
            (3,"Nottm Forest",    29,17,6,6, 48,35,13,57,40.0,36.2),
            (4,"Chelsea",         29,15,8,6, 62,44,18,53,57.1,44.3),
            (5,"Man City",        29,14,7,8, 56,41,15,49,62.1,43.8),
            (6,"Newcastle",       29,13,6,10,51,39,12,45,48.7,38.4),
            (7,"Aston Villa",     29,12,7,10,54,47, 7,43,54.2,50.1),
            (8,"Tottenham",       29,11,5,13,46,55,-9,38,48.0,50.0),
            (9,"Man United",      29,10,5,14,35,46,-11,35,38.5,52.2),
            (10,"Brighton",       29, 9,8,12,46,46, 0,35,47.5,47.0),
            (11,"Fulham",         29, 9,7,13,40,46,-6,34,37.1,45.4),
            (12,"Brentford",      29, 9,5,15,43,55,-12,32,40.3,54.0),
            (13,"West Ham",       29, 8,6,15,35,49,-14,30,36.0,48.3),
            (14,"Everton",        29, 7,8,14,29,43,-14,29,31.1,42.2),
            (15,"Crystal Palace", 29, 6,9,14,29,45,-16,27,30.0,42.0),
            (16,"Wolves",         29, 6,5,18,38,62,-24,23,36.2,58.0),
            (17,"Ipswich",        29, 4,8,17,28,57,-29,20,30.0,57.0),
            (18,"Leicester",      29, 4,7,18,32,66,-34,19,29.5,62.2),
            (19,"Southampton",    29, 2,6,21,22,68,-46,12,22.0,64.0),
            (20,"Sunderland",     29, 2,4,23,20,71,-51,10,20.0,66.0),
        ]
        return pd.DataFrame(data, columns=self.COLS+self.EXTRA_COLS)


# ══════════════════════════ EXCEL EXPORTER ════════════════════════════════
class ExcelExporter:
    PL_PURPLE = "38003C"; PL_GREEN = "00FF85"
    LIGHT_ROW = "F5F0F6"; DARK_ROW  = "EAE0EB"

    def export(self, df, path):
        wb = Workbook(); ws = wb.active; ws.title = "PL Standings"
        hf   = PatternFill("solid", fgColor=self.PL_PURPLE)
        hfnt = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
        ha   = Alignment(horizontal="center", vertical="center")
        thin = Side(style="thin", color="D0C0D4")
        brd  = Border(left=thin, right=thin, top=thin, bottom=thin)
        cols = list(df.columns)
        for ci, col in enumerate(cols, 1):
            cell = ws.cell(row=1, column=ci, value=col)
            cell.fill=hf; cell.font=hfnt; cell.alignment=ha
        for ri, row in df.iterrows():
            fc = self.LIGHT_ROW if ri%2==0 else self.DARK_ROW
            rf = PatternFill("solid", fgColor=fc)
            rank = int(row.get("Rank", ri+1))
            rc = "1A73E8" if rank<=4 else "FF9500" if rank<=6 else "E53935" if rank>=18 else "38003C"
            for ci, col in enumerate(cols, 1):
                v = row[col]; v = "" if pd.isna(v) else v
                cell = ws.cell(row=ri+2, column=ci, value=v)
                cell.fill=rf; cell.border=brd
                cell.alignment = Alignment(horizontal="left" if col=="Team" else "center", vertical="center")
                if col=="Rank": cell.font=Font(bold=True, color=rc, name="Calibri")
                elif col=="Team": cell.font=Font(bold=True, name="Calibri")
                else: cell.font=Font(name="Calibri")
        for ci, col in enumerate(cols, 1):
            mw = max(len(str(col)), *(len(str(df.iloc[r][col])) for r in range(len(df))))
            ws.column_dimensions[get_column_letter(ci)].width = mw+4
        ws.row_dimensions[1].height = 22
        # Chart sheet
        cws = wb.create_sheet("Points Chart")
        cws["A1"]="Team"; cws["B1"]="Points"
        for ri, row in df.iterrows():
            cws.cell(row=ri+2, column=1, value=row["Team"])
            cws.cell(row=ri+2, column=2, value=row["Pts"])
        chart = BarChart(); chart.type="bar"
        chart.title="Premier League 2024-25 — Points"; chart.style=10
        chart.y_axis.title="Points"; chart.x_axis.title="Team"
        dr = Reference(cws, min_col=2, min_row=1, max_row=len(df)+1)
        lr = Reference(cws, min_col=1, min_row=2, max_row=len(df)+1)
        chart.add_data(dr, titles_from_data=True); chart.set_categories(lr)
        chart.width=28; chart.height=16; cws.add_chart(chart, "D2")
        wb.save(path)


# ══════════════════════════════ MAIN APP ═══════════════════════════════════
class PLStatsApp(tk.Tk):

    def __init__(self):
        super().__init__()
        self.title("⚽ Premier League 2024-25  |  Live Standings + Sort Visualizer")
        self.geometry("1400x860")
        self.minsize(1100, 720)
        self.configure(bg=BG_DARK)

        self.df          = pd.DataFrame()
        self.filtered_df = pd.DataFrame()
        self.data_source = "—"
        self.is_loading  = False
        self._sort_col   = "Pts"
        self._sort_asc   = False
        self._search_var = tk.StringVar()
        self._search_var.trace_add("write", self._on_search)

        # Sort-viz state
        self._viz_sorting  = False
        self._viz_paused   = False
        self._viz_gen      = None
        self._viz_after_id = None
        self._viz_comparisons = 0
        self._viz_swaps       = 0
        self._viz_start_time  = 0
        self._viz_sorted_idx  = set()
        self._viz_teams       = []   # team names in original order for labels

        self._fetcher  = DataFetcher()
        self._exporter = ExcelExporter()

        self._build_ui()
        self._fetch_data()

    # ══════════════════════ UI CONSTRUCTION ═══════════════════════════
    def _build_ui(self):
        self._build_titlebar()
        # Notebook (tabs)
        nb_style = ttk.Style()
        nb_style.theme_use("clam")
        nb_style.configure("PL.TNotebook",         background=BG_DARK,   borderwidth=0)
        nb_style.configure("PL.TNotebook.Tab",      background=BG_PANEL,  foreground=TEXT_DIM,
                           font=("Segoe UI",10,"bold"), padding=[16,8])
        nb_style.map("PL.TNotebook.Tab",
                     background=[("selected", ACCENT)],
                     foreground=[("selected", ACCENT2)])

        self.nb = ttk.Notebook(self, style="PL.TNotebook")
        self.nb.pack(fill="both", expand=True, padx=8, pady=(4,4))

        # Tab 1 – Standings
        tab1 = tk.Frame(self.nb, bg=BG_DARK)
        self.nb.add(tab1, text="📋  Live Standings")
        self._build_standings_tab(tab1)

        # Tab 2 – Sort Visualizer
        tab2 = tk.Frame(self.nb, bg=BG_DARK)
        self.nb.add(tab2, text="⚡  Sort Visualizer")
        self._build_viz_tab(tab2)

        self._build_statusbar()

    # ── Title bar ─────────────────────────────────────────────────────
    def _build_titlebar(self):
        bar = tk.Frame(self, bg=ACCENT, height=58)
        bar.pack(fill="x"); bar.pack_propagate(False)
        tk.Label(bar, text="⚽", font=("Segoe UI Emoji",22), fg=ACCENT2, bg=ACCENT
                 ).pack(side="left", padx=(16,4), pady=8)
        tk.Label(bar, text="PREMIER LEAGUE", font=("Segoe UI",17,"bold"),
                 fg=ACCENT2, bg=ACCENT).pack(side="left", pady=8)
        tk.Label(bar, text="  2024–25  Live Standings + Sort Visualizer",
                 font=("Segoe UI",12), fg="#C8A2CB", bg=ACCENT).pack(side="left", pady=8)

        def btn(text, cmd, bg, fg="#000000"):
            b = tk.Button(bar, text=text, command=cmd, bg=bg, fg=fg,
                          font=("Segoe UI",10,"bold"), relief="flat", cursor="hand2",
                          padx=14, pady=6, activebackground=bg, bd=0, highlightthickness=0)
            b.pack(side="right", padx=6, pady=10); return b

        btn("📊  Export to Excel", self._export_excel, ACCENT2)
        self.refresh_btn = btn("🔄  Refresh", self._fetch_data, "#4A0050", "#FFFFFF")

    # ══════════════════════ TAB 1: STANDINGS ══════════════════════════
    def _build_standings_tab(self, parent):
        main = tk.Frame(parent, bg=BG_DARK)
        main.pack(fill="both", expand=True, padx=6, pady=4)
        left = tk.Frame(main, bg=BG_DARK)
        left.pack(side="left", fill="both", expand=True)
        right = tk.Frame(main, bg=BG_PANEL, width=420)
        right.pack(side="right", fill="y", padx=(8,0))
        right.pack_propagate(False)
        self._build_stat_cards(left)
        self._build_search(left)
        self._build_table(left)
        self._build_points_chart(right)

    def _build_stat_cards(self, parent):
        self.cards_frame = tk.Frame(parent, bg=BG_DARK)
        self.cards_frame.pack(fill="x", pady=(6,8))
        cards = [
            ("🏆  LEADER",       "leader_val",  "—", TEXT_GOLD),
            ("⚽  TOTAL GOALS",  "goals_val",   "—", ACCENT2),
            ("🏅  MOST WINS",    "wins_val",    "—", "#64B5F6"),
            ("📊  AVG POINTS",   "avgpts_val",  "—", "#CE93D8"),
            ("🔒  BEST DEFENCE", "def_val",     "—", "#A5D6A7"),
        ]
        for title, attr, default, color in cards:
            card = tk.Frame(self.cards_frame, bg=BG_CARD, padx=14, pady=8)
            card.pack(side="left", padx=4, fill="y", expand=True)
            tk.Label(card, text=title, font=("Segoe UI",8,"bold"),
                     fg=TEXT_DIM, bg=BG_CARD).pack(anchor="w")
            lbl = tk.Label(card, text=default, font=("Consolas",13,"bold"),
                           fg=color, bg=BG_CARD)
            lbl.pack(anchor="w")
            setattr(self, attr, lbl)

    def _update_stat_cards(self):
        if self.df.empty: return
        df = self.df
        if "Pts" in df.columns:
            self.leader_val.config(text=df.loc[df["Pts"].idxmax(),"Team"])
        if "GF" in df.columns:
            self.goals_val.config(text=f"{int(df['GF'].sum()):,}")
        if "W" in df.columns:
            bw = df.loc[df["W"].idxmax()]
            self.wins_val.config(text=f"{bw['Team']} ({int(bw['W'])}W)")
        if "Pts" in df.columns:
            self.avgpts_val.config(text=f"{df['Pts'].mean():.1f}")
        if "GA" in df.columns:
            bd = df.loc[df["GA"].idxmin()]
            self.def_val.config(text=f"{bd['Team']} ({int(bd['GA'])}GA)")

    def _build_search(self, parent):
        sf = tk.Frame(parent, bg=BG_PANEL, pady=6)
        sf.pack(fill="x", pady=(0,6))
        tk.Label(sf, text="🔍  Search team:", font=FONT_BODY,
                 fg=TEXT_DIM, bg=BG_PANEL).pack(side="left", padx=10)
        entry = tk.Entry(sf, textvariable=self._search_var, font=FONT_BODY,
                         bg=BG_CARD, fg=TEXT_MAIN, insertbackground=TEXT_MAIN,
                         relief="flat", bd=4, highlightthickness=1,
                         highlightcolor=ACCENT2, highlightbackground=BG_CARD)
        entry.pack(side="left", padx=4, ipady=4, fill="x", expand=True)
        tk.Button(sf, text="✕", command=lambda: self._search_var.set(""),
                  bg=BG_CARD, fg=TEXT_DIM, font=FONT_SMALL, relief="flat",
                  cursor="hand2", bd=0, padx=8).pack(side="left", padx=(0,10))

    def _build_table(self, parent):
        sty = ttk.Style()
        sty.configure("PL.Treeview", background=BG_CARD, foreground=TEXT_MAIN,
                       fieldbackground=BG_CARD, rowheight=28, font=FONT_BODY, borderwidth=0)
        sty.configure("PL.Treeview.Heading", background=ACCENT, foreground=ACCENT2,
                       font=("Segoe UI",10,"bold"), relief="flat", borderwidth=0)
        sty.map("PL.Treeview",
                background=[("selected","#4A0050")], foreground=[("selected",ACCENT2)])
        cols = ["Rank","Team","MP","W","D","L","GF","GA","GD","Pts","xG","xGA"]
        widths = [45, 160, 45, 45, 45, 45, 50, 50, 55, 55, 60, 60]
        self.tree = ttk.Treeview(parent, columns=cols, show="headings", style="PL.Treeview")
        for col, w in zip(cols, widths):
            self.tree.heading(col, text=col, command=lambda c=col: self._on_header_click(c))
            self.tree.column(col, width=w, anchor="w" if col=="Team" else "center", minwidth=30)
        self.tree.tag_configure("cl",     foreground="#90CAF9", background="#0D1B35")
        self.tree.tag_configure("el",     foreground="#FFCC80", background="#1E1500")
        self.tree.tag_configure("rel",    foreground="#EF9A9A", background="#1E0505")
        self.tree.tag_configure("normal", foreground=TEXT_MAIN, background=BG_CARD)
        self.tree.tag_configure("alt",    foreground=TEXT_MAIN, background=BG_ROW_ALT)
        vsb = ttk.Scrollbar(parent, orient="vertical",   command=self.tree.yview)
        hsb = ttk.Scrollbar(parent, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        tf = tk.Frame(parent, bg=BG_DARK)
        tf.pack(fill="both", expand=True)
        self.tree.grid(in_=tf, row=0, column=0, sticky="nsew")
        vsb.grid(in_=tf, row=0, column=1, sticky="ns")
        hsb.grid(in_=tf, row=1, column=0, sticky="ew")
        tf.grid_rowconfigure(0, weight=1); tf.grid_columnconfigure(0, weight=1)

    def _build_points_chart(self, parent):
        tk.Label(parent, text="📈  POINTS BY TEAM", font=("Segoe UI",10,"bold"),
                 fg=ACCENT2, bg=BG_PANEL).pack(anchor="w", padx=14, pady=(10,2))
        self.fig_pts = Figure(figsize=(4.2,7), facecolor=BG_PANEL)
        self.ax_pts  = self.fig_pts.add_subplot(111)
        self._style_pts_ax()
        self.canvas_pts = FigureCanvasTkAgg(self.fig_pts, master=parent)
        self.canvas_pts.get_tk_widget().pack(fill="both", expand=True, padx=4, pady=(0,4))
        self.canvas_pts.draw()
        lf = tk.Frame(parent, bg=BG_PANEL)
        lf.pack(fill="x", padx=10, pady=(0,8))
        for label, color in [("CL",CL_COLOR),("EL",EL_COLOR),("Mid",MID_COLOR),("Rel",REL_COLOR)]:
            f = tk.Frame(lf, bg=BG_PANEL); f.pack(side="left", padx=6)
            tk.Label(f, text="■", fg=color, bg=BG_PANEL, font=("Segoe UI",14)).pack(side="left")
            tk.Label(f, text=label, fg=TEXT_DIM, bg=BG_PANEL, font=FONT_SMALL).pack(side="left")

    def _style_pts_ax(self):
        self.ax_pts.set_facecolor(BG_PANEL)
        self.ax_pts.tick_params(colors=TEXT_DIM, labelsize=8)
        for sp in self.ax_pts.spines.values(): sp.set_visible(False)
        self.fig_pts.tight_layout(pad=1.2)

    def _update_points_chart(self):
        if self.filtered_df.empty: return
        df = self.filtered_df.sort_values("Pts", ascending=True) if "Pts" in self.filtered_df.columns else self.filtered_df
        self.ax_pts.clear(); self._style_pts_ax()
        teams, pts = df["Team"].tolist(), df["Pts"].tolist()
        n = len(teams)
        colors = []
        for _, row in df.iterrows():
            rk = row.get("Rank",99)
            colors.append(CL_COLOR if rk<=4 else EL_COLOR if rk<=6 else REL_COLOR if rk>=18 else MID_COLOR)
        bars = self.ax_pts.barh(range(n), pts, color=colors, edgecolor="none", height=0.7)
        self.ax_pts.set_yticks(range(n)); self.ax_pts.set_yticklabels(teams, fontsize=8, color=TEXT_MAIN)
        self.ax_pts.set_xlabel("Points", color=TEXT_DIM, fontsize=8)
        self.ax_pts.tick_params(axis="x", colors=TEXT_DIM, labelsize=7)
        for bar, p in zip(bars, pts):
            if not np.isnan(p):
                self.ax_pts.text(bar.get_width()+0.3, bar.get_y()+bar.get_height()/2,
                                 f"{int(p)}", va="center", ha="left", fontsize=7, color=TEXT_MAIN)
        self.canvas_pts.draw_idle()

    # ══════════════════════ TAB 2: SORT VISUALIZER ════════════════════
    def _build_viz_tab(self, parent):
        # Left: controls + stats strip + chart
        left = tk.Frame(parent, bg=BG_DARK)
        left.pack(side="left", fill="both", expand=True, padx=(4,0), pady=4)

        # Right: sidebar
        side = tk.Frame(parent, bg=BG_PANEL, width=250)
        side.pack(side="right", fill="y", padx=(8,4), pady=4)
        side.pack_propagate(False)
        self._build_viz_sidebar(side)

        # Stats strip
        sstrip = tk.Frame(left, bg=BG_PANEL, height=58)
        sstrip.pack(fill="x", pady=(0,6)); sstrip.pack_propagate(False)
        self._build_viz_stats(sstrip)

        # Chart area
        cframe = tk.Frame(left, bg=BG_DARK)
        cframe.pack(fill="both", expand=True)
        self._build_viz_chart(cframe)

        # Legend
        lf = tk.Frame(left, bg=BG_PANEL, height=30)
        lf.pack(fill="x", pady=(4,0)); lf.pack_propagate(False)
        legend_items = [("Default",COL_DEFAULT),("Comparing",COL_COMPARE),
                        ("Swapping",COL_SWAP),  ("Sorted",COL_SORTED)]
        for label, color in legend_items:
            f = tk.Frame(lf, bg=BG_PANEL); f.pack(side="left", padx=12)
            tk.Label(f, text="■", fg=color, bg=BG_PANEL, font=("Segoe UI",13)).pack(side="left")
            tk.Label(f, text=label, fg=TEXT_DIM, bg=BG_PANEL, font=FONT_SMALL).pack(side="left")

    def _build_viz_sidebar(self, parent):
        pad = dict(padx=14, pady=5)

        def section(text):
            tk.Label(parent, text=text, font=("Segoe UI",9,"bold"),
                     fg=ACCENT2, bg=BG_PANEL).pack(anchor="w", **pad)

        section("ALGORITHM")
        self.viz_algo_var = tk.StringVar(value="Bubble Sort")
        cb = ttk.Combobox(parent, textvariable=self.viz_algo_var,
                          values=list(ALGORITHMS.keys()), state="readonly", font=FONT_BODY)
        cb.pack(fill="x", padx=14, pady=(0,8))
        cb.bind("<<ComboboxSelected>>", lambda e: self._update_viz_complexity())

        section("SORT BY COLUMN")
        self.viz_col_var = tk.StringVar(value="Pts")
        col_cb = ttk.Combobox(parent, textvariable=self.viz_col_var,
                              values=["Pts","GF","GA","GD","W","D","L","MP"],
                              state="readonly", font=FONT_BODY)
        col_cb.pack(fill="x", padx=14, pady=(0,8))

        section("SPEED")
        self.viz_speed_var = tk.IntVar(value=60)
        self.viz_speed_lbl = tk.Label(parent, text="60 ms", font=FONT_MONO,
                                      fg=TEXT_MAIN, bg=BG_PANEL)
        self.viz_speed_lbl.pack(anchor="e", padx=18)
        ttk.Scale(parent, from_=1, to=300, variable=self.viz_speed_var, orient="horizontal",
                  command=lambda v: self.viz_speed_lbl.config(text=f"{int(float(v))} ms")
                  ).pack(fill="x", padx=14, pady=(0,14))

        # Buttons
        def vbtn(text, cmd, color):
            b = tk.Button(parent, text=text, command=cmd, bg=color, fg="white",
                          font=("Segoe UI",10,"bold"), relief="flat", cursor="hand2",
                          pady=8, activebackground=color, bd=0, highlightthickness=0)
            b.pack(fill="x", padx=14, pady=4); return b

        self.viz_start_btn  = vbtn("▶  Start Visualizer", self._viz_start,  ACCENT)
        self.viz_pause_btn  = vbtn("⏸  Pause / Resume",   self._viz_toggle_pause, "#546E7A")
        vbtn("⏹  Reset",              self._viz_reset,   "#E53935")

        # Complexity card
        tk.Frame(parent, bg=BG_CARD, height=2).pack(fill="x", padx=14, pady=(14,8))
        tk.Label(parent, text="COMPLEXITY", font=("Segoe UI",9,"bold"),
                 fg=ACCENT2, bg=BG_PANEL).pack(anchor="w", padx=14)
        cxf = tk.Frame(parent, bg=BG_CARD); cxf.pack(fill="x", padx=14, pady=6)
        self.cx_best  = self._cx_row(cxf, "Best",  "—")
        self.cx_worst = self._cx_row(cxf, "Worst", "—")
        self.cx_space = self._cx_row(cxf, "Space", "—")

        # Explain label — must be created BEFORE calling _update_viz_complexity
        tk.Frame(parent, bg=BG_CARD, height=2).pack(fill="x", padx=14, pady=(10,8))
        self.cx_explain = tk.Label(parent, text="", font=("Segoe UI",8), fg=TEXT_DIM,
                                   bg=BG_PANEL, wraplength=210, justify="left")
        self.cx_explain.pack(anchor="w", padx=14)
        self._update_viz_complexity()

    def _cx_row(self, parent, label, val):
        row = tk.Frame(parent, bg=BG_CARD); row.pack(fill="x", padx=6, pady=2)
        tk.Label(row, text=label+":", font=FONT_SMALL, fg=TEXT_DIM, bg=BG_CARD,
                 width=6, anchor="w").pack(side="left")
        v = tk.Label(row, text=val, font=("Consolas",10,"bold"),
                     fg=COL_COMPARE, bg=BG_CARD)
        v.pack(side="left"); return v

    def _update_viz_complexity(self, *_):
        algo = self.viz_algo_var.get()
        best, worst, space = COMPLEXITY.get(algo, ("—","—","—"))
        if not hasattr(self, "cx_best"): return
        self.cx_best.config(text=best)
        self.cx_worst.config(text=worst)
        self.cx_space.config(text=space)
        explains = {
            "Bubble Sort":    "Repeatedly swaps adjacent elements if in wrong order.",
            "Selection Sort": "Finds the minimum and places it at the start each pass.",
            "Insertion Sort": "Inserts each element into its correct position.",
            "Merge Sort":     "Divides array in half, sorts each, then merges.",
            "Quick Sort":     "Picks a pivot, partitions around it, recurses.",
            "Heap Sort":      "Builds a max-heap, extracts maximum repeatedly.",
        }
        self.cx_explain.config(text=explains.get(algo,""))

    def _build_viz_stats(self, parent):
        stats = [
            ("⚙  COMPARISONS", "viz_cmp_lbl",  "0"),
            ("🔄  SWAPS",       "viz_swp_lbl",  "0"),
            ("⏱  TIME",        "viz_time_lbl", "0.000 s"),
            ("📐  TEAMS",       "viz_sz_lbl",   "—"),
            ("🏷  ALGORITHM",   "viz_alg_lbl",  "—"),
        ]
        for title, attr, default in stats:
            card = tk.Frame(parent, bg=BG_CARD, padx=14, pady=6)
            card.pack(side="left", padx=6, pady=6, fill="y")
            tk.Label(card, text=title, font=("Segoe UI",8,"bold"),
                     fg=TEXT_DIM, bg=BG_CARD).pack(anchor="w")
            lbl = tk.Label(card, text=default, font=("Consolas",13,"bold"),
                           fg=TEXT_MAIN, bg=BG_CARD)
            lbl.pack(anchor="w")
            setattr(self, attr, lbl)

    def _build_viz_chart(self, parent):
        self.fig_viz = Figure(figsize=(9,4.5), facecolor=BG_DARK)
        self.ax_viz  = self.fig_viz.add_subplot(111)
        self._style_viz_ax()
        self.canvas_viz = FigureCanvasTkAgg(self.fig_viz, master=parent)
        self.canvas_viz.get_tk_widget().pack(fill="both", expand=True)
        self.canvas_viz.draw()

    def _style_viz_ax(self):
        self.ax_viz.set_facecolor(BG_DARK)
        self.ax_viz.tick_params(colors=TEXT_DIM, labelsize=7)
        for sp in ("top","right"): self.ax_viz.spines[sp].set_visible(False)
        self.ax_viz.spines["bottom"].set_color(BG_CARD)
        self.ax_viz.spines["left"].set_color(BG_CARD)
        self.ax_viz.set_ylabel("Value", color=TEXT_DIM, fontsize=8)

    # ── Viz drawing ───────────────────────────────────────────────────
    def _viz_draw(self, values, teams, highlight, state, col_name):
        self.ax_viz.clear(); self._style_viz_ax()
        n = len(values)
        colors = []
        for i in range(n):
            if i in self._viz_sorted_idx or state=="done":
                colors.append(COL_SORTED)
            elif i in highlight:
                colors.append(COL_SWAP if state=="swap" else
                              COL_COMPARE if state=="compare" else COL_SORTED)
            else:
                colors.append(COL_DEFAULT)

        bars = self.ax_viz.bar(range(n), values, color=colors, edgecolor="none", width=0.85)
        self.ax_viz.set_xlim(-0.5, n-0.5)
        self.ax_viz.set_ylim(0, max(values)*1.12 if n>0 else 100)
        self.ax_viz.set_xticks(range(n))
        self.ax_viz.set_xticklabels(teams, rotation=45, ha="right", fontsize=6.5, color=TEXT_DIM)
        self.ax_viz.set_title(
            f"{self.viz_algo_var.get()}  —  Sorting by: {col_name}",
            color=TEXT_MAIN, fontsize=10, pad=6
        )
        # Value labels on bars
        for bar, v in zip(bars, values):
            if bar.get_height() > 0:
                self.ax_viz.text(bar.get_x()+bar.get_width()/2, bar.get_height()+0.5,
                                 str(int(v)), ha="center", va="bottom",
                                 fontsize=6, color=TEXT_MAIN)
        self.fig_viz.tight_layout()
        self.canvas_viz.draw_idle()

    # ── Viz controls ─────────────────────────────────────────────────
    def _viz_start(self):
        if self._viz_sorting: return
        if self.df.empty:
            messagebox.showwarning("No Data", "Please wait for data to load first.")
            return
        col = self.viz_col_var.get()
        if col not in self.df.columns:
            messagebox.showerror("Error", f"Column '{col}' not found in loaded data.")
            return
        # Build working array from real PL data
        working = self.df[["Team", col]].dropna().copy()
        self._viz_teams  = working["Team"].tolist()
        values_arr       = working[col].astype(float).tolist()

        self._viz_sorting      = True
        self._viz_paused       = False
        self._viz_comparisons  = 0
        self._viz_swaps        = 0
        self._viz_start_time   = time.perf_counter()
        self._viz_sorted_idx   = set()

        algo = self.viz_algo_var.get()
        self.viz_alg_lbl.config(text=algo)
        self.viz_sz_lbl.config(text=str(len(values_arr)))
        self.viz_cmp_lbl.config(text="0")
        self.viz_swp_lbl.config(text="0")
        self.viz_time_lbl.config(text="0.000 s")

        self._viz_gen = ALGORITHMS[algo](values_arr)
        self._viz_step(values_arr, col)

    def _viz_step(self, original_values, col):
        if not self._viz_sorting: return
        if self._viz_paused:
            self._viz_after_id = self.after(100, lambda: self._viz_step(original_values, col))
            return
        try:
            arr, highlight, state, cmp_inc, swp_inc = next(self._viz_gen)
            self._viz_comparisons += cmp_inc
            self._viz_swaps       += swp_inc
            if state == "sorted_mark":
                for idx in highlight: self._viz_sorted_idx.add(idx)
            elapsed = time.perf_counter() - self._viz_start_time
            self.viz_cmp_lbl.config(text=f"{self._viz_comparisons:,}")
            self.viz_swp_lbl.config(text=f"{self._viz_swaps:,}")
            self.viz_time_lbl.config(text=f"{elapsed:.3f} s")
            # Build team labels in current permutation order
            n = len(arr)
            # Sort teams by current index permutation (track by value position changes)
            self._viz_draw(arr, self._viz_teams, highlight, state, col)
            delay = max(1, int(self.viz_speed_var.get()))
            self._viz_after_id = self.after(delay, lambda: self._viz_step(original_values, col))
        except StopIteration:
            self._viz_done(col)

    def _viz_done(self, col):
        self._viz_sorting = False
        elapsed = time.perf_counter() - self._viz_start_time
        self.viz_time_lbl.config(text=f"{elapsed:.3f} s")
        self._viz_sorted_idx = set(range(len(self._viz_teams)))
        # Show final sorted state
        sorted_df = self.df[["Team",col]].dropna().sort_values(col, ascending=True)
        self._viz_draw(sorted_df[col].tolist(), sorted_df["Team"].tolist(),
                       list(range(len(sorted_df))), "done", col)

    def _viz_toggle_pause(self):
        if not self._viz_sorting: return
        self._viz_paused = not self._viz_paused
        self.viz_pause_btn.config(
            text="▶  Resume" if self._viz_paused else "⏸  Pause / Resume"
        )

    def _viz_reset(self):
        if self._viz_after_id: self.after_cancel(self._viz_after_id)
        self._viz_sorting = False; self._viz_paused = False
        self._viz_comparisons = self._viz_swaps = 0
        self._viz_sorted_idx = set()
        self.viz_cmp_lbl.config(text="0"); self.viz_swp_lbl.config(text="0")
        self.viz_time_lbl.config(text="0.000 s")
        self.viz_pause_btn.config(text="⏸  Pause / Resume")
        self.ax_viz.clear(); self._style_viz_ax(); self.canvas_viz.draw_idle()

    # ══════════════════════ DATA LOADING ══════════════════════════════
    def _fetch_data(self):
        if self.is_loading: return
        self.is_loading = True
        self.refresh_btn.config(state="disabled", text="⏳  Loading…")
        self.status_lbl.config(text="Fetching live data…")
        self._start_spinner()
        threading.Thread(target=self._fetch_thread, daemon=True).start()

    def _fetch_thread(self):
        df, source = self._fetcher.fetch()
        self.after(0, lambda: self._on_data_ready(df, source))

    def _on_data_ready(self, df, source):
        self.is_loading = False; self.data_source = source
        self._stop_spinner()
        self.refresh_btn.config(state="normal", text="🔄  Refresh")
        if df.empty:
            messagebox.showerror("Error", "Could not fetch data.")
            self.status_lbl.config(text="Failed to load data."); return
        self.df = df; self._apply_sort()
        now = datetime.now().strftime("%H:%M:%S")
        self.status_lbl.config(text=f"Last updated: {now}  |  Source: {source}  |  Season: 2024-25")

    # ══════════════════════ SORTING (STANDINGS) ════════════════════════
    def _on_header_click(self, col):
        if self._sort_col == col: self._sort_asc = not self._sort_asc
        else: self._sort_col = col; self._sort_asc = (col=="Team")
        self._apply_sort()

    def _apply_sort(self):
        if self.df.empty: return
        col, asc = self._sort_col, self._sort_asc
        if col in self.df.columns:
            self.df = self.df.sort_values(col, ascending=asc, ignore_index=True)
            self.df["Rank"] = range(1, len(self.df)+1)
        self._apply_filter()

    def _on_search(self, *_): self._apply_filter()

    def _apply_filter(self):
        query = self._search_var.get().strip().lower()
        self.filtered_df = (
            self.df[self.df["Team"].str.lower().str.contains(query)]
            if query and "Team" in self.df.columns else self.df.copy()
        )
        self._refresh_table(); self._update_stat_cards(); self._update_points_chart()
        self._update_tree_headings()
        self.rows_lbl.config(text=f"Showing {len(self.filtered_df)} / {len(self.df)} teams")

    def _update_tree_headings(self):
        for col in ["Rank","Team","MP","W","D","L","GF","GA","GD","Pts","xG","xGA"]:
            ind = (" ▲" if self._sort_asc else " ▼") if self._sort_col==col else ""
            self.tree.heading(col, text=col+ind, command=lambda c=col: self._on_header_click(c))

    def _refresh_table(self):
        self.tree.delete(*self.tree.get_children())
        cols = ["Rank","Team","MP","W","D","L","GF","GA","GD","Pts","xG","xGA"]
        for i, row in self.filtered_df.iterrows():
            vals = []
            for col in cols:
                if col not in self.filtered_df.columns: vals.append("—"); continue
                v = row[col]
                if pd.isna(v): vals.append("—")
                elif col in ("xG","xGA"): vals.append(f"{v:.1f}")
                elif col=="Team": vals.append(str(v))
                else: vals.append(int(v))
            rank = row.get("Rank", i+1)
            tag = "cl" if rank<=4 else "el" if rank<=6 else "rel" if rank>=18 else ("normal" if i%2==0 else "alt")
            self.tree.insert("", "end", values=vals, tags=(tag,))

    # ══════════════════════ EXCEL EXPORT ══════════════════════════════
    def _export_excel(self):
        if self.df.empty:
            messagebox.showwarning("No Data","Load data before exporting."); return
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel Workbook","*.xlsx")],
            initialfile="PL_Standings_2024_25.xlsx",
            title="Save Premier League Standings"
        )
        if not path: return
        try:
            edf = self.df.copy()
            for col in ("xG","xGA"):
                if col in edf.columns:
                    edf[col] = edf[col].apply(lambda x: round(x,1) if not pd.isna(x) else "")
            self._exporter.export(edf, path)
            messagebox.showinfo("✅ Exported", f"Saved successfully!\n\n{path}")
        except Exception as e:
            messagebox.showerror("Export Error", str(e))

    # ══════════════════════ STATUS BAR ════════════════════════════════
    def _build_statusbar(self):
        bar = tk.Frame(self, bg=BG_PANEL, height=28)
        bar.pack(fill="x", side="bottom"); bar.pack_propagate(False)
        self.status_lbl = tk.Label(bar, text="Loading data…",
                                   font=FONT_SMALL, fg=TEXT_DIM, bg=BG_PANEL)
        self.status_lbl.pack(side="left", padx=12, pady=4)
        self.spinner_lbl = tk.Label(bar, text="", font=FONT_SMALL,
                                    fg=ACCENT2, bg=BG_PANEL)
        self.spinner_lbl.pack(side="left", padx=4)
        self.rows_lbl = tk.Label(bar, text="", font=FONT_SMALL,
                                  fg=TEXT_DIM, bg=BG_PANEL)
        self.rows_lbl.pack(side="right", padx=12)

    # ══════════════════════ SPINNER ═══════════════════════════════════
    _frames = ["⠋","⠙","⠹","⠸","⠼","⠴","⠦","⠧","⠇","⠏"]
    _fidx   = 0; _fjob = None

    def _start_spinner(self):
        self._fidx = 0; self._spin()

    def _spin(self):
        self.spinner_lbl.config(text=self._frames[self._fidx % len(self._frames)])
        self._fidx += 1
        self._fjob = self.after(80, self._spin)

    def _stop_spinner(self):
        if self._fjob: self.after_cancel(self._fjob); self._fjob = None
        self.spinner_lbl.config(text="✔")


# ─────────────────────────────── ENTRY POINT ───────────────────────────────
if __name__ == "__main__":
    app = PLStatsApp()
    app.mainloop()
